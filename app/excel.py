import pandas as pd
import numpy as np
from openpyxl import load_workbook
import re

DEFAULT_FILENAME = 'Raspisanie_2_sem_2022_23_uch_g_2023_13_02_23.xlsx'
DEFAULT_SHEET = 'Расписание'

# variables
groups_row = 16
day_column = 1
time_column = 2

find_week = r'\(\d+-\d+ [а-я]+\)|н\/н|ч\/н'
find_teacher = r'[А-ЯЁ][а-яё]+ [А-Я].[А-Я].'
find_street_or_build = r'\([А-Яа-я]+\.[0-9 ]+\)'
find_classroom = r'ауд.[0-9А-ЯЁа-яё ]+'

def read_excel(filename, sheet_name=DEFAULT_SHEET):
    wb = load_workbook(filename=filename)
    timetable = wb[sheet_name]
    return (wb, timetable)

def save_changes(workbook, filename):
    workbook.save(filename)

def get_merged_cells(sheet):
    return sheet.merged_cells.ranges

def get_cell(sheet, rrange):
    return sheet.cell(row=rrange.bounds[1], column=rrange.bounds[0]).value

def get_cell(sheet, column, row):
    return sheet.cell(row=row, column=column).value

def fill_merged_cells(sheet, cell_ranges):

    bounds = [rrange.bounds for rrange in cell_ranges]

    for merge in list(sheet.merged_cells):
        sheet.unmerge_cells(range_string=str(merge))

    for bound in bounds:
        for j in range(bound[0], bound[2]+1):
            for i in range(bound[1], bound[3]+1):
                sheet.cell(row=i, column=j).value = sheet.cell(row=bound[1], column=bound[0]).value

def get_bottom_border(sheet):
    cur_row = groups_row
    while sheet.cell(row=cur_row, column=day_column).value is not None:
        cur_row += 1
    
    return cur_row - 1

def get_end_border(sheet):
    cur_column = day_column
    while sheet.cell(row=groups_row, column=cur_column).value is not None:
        cur_column += 1
    
    return cur_column - 1

def get_groupp():
    wb, timetable = read_excel(DEFAULT_FILENAME)
    groupp = []
    fill_merged_cells(timetable, get_merged_cells(timetable))
    bottom, end = get_bottom_border(timetable), get_end_border(timetable)
    for i in range(time_column + 1, end + 1, 2):
        if(timetable.cell(row=groups_row, column=i).value.split(' ')[0] != "МАГИСТРАТУРА"):
            groupp.append(timetable.cell(row=groups_row, column=i).value.split(' ')[0])

    return groupp


def get_timetable():
    wb, timetable = read_excel(DEFAULT_FILENAME)
    groups_timetable = []

    fill_merged_cells(timetable, get_merged_cells(timetable))
    bottom, end = get_bottom_border(timetable), get_end_border(timetable)

    for i in range(time_column + 1, end + 1, 2):
        gr_num = timetable.cell(row=groups_row, column=i).value.split(' ')[0]
        for j in range(groups_row + 1, bottom + 1, 2):
            one_lesson = timetable.cell(row=j, column=i).value
            if(one_lesson is not None):
                one_lesson.strip()
                if(one_lesson == "понедельник" or one_lesson == "вт" or one_lesson == "ср" or one_lesson == "чт" or one_lesson == "пт" or one_lesson == "сб" or one_lesson == "вторник" or one_lesson == "среда" or one_lesson == "четверг" or one_lesson == "пятница" or one_lesson == "суббота" or one_lesson == '') :
                    continue
                else:
                    match_week = re.findall(find_week, one_lesson)
                    match_teacher = re.findall(find_teacher, one_lesson)
                    match_search_teacher = re.search(find_teacher, one_lesson)
                    match_address = re.findall(find_street_or_build, one_lesson)
                    match_search_address = re.search(find_classroom, one_lesson)
                    match_search_index_start_subject = match_search_address.start() if match_search_teacher is None else match_search_teacher.start()
                    match_firsct_close_sc = re.search(r'\)', one_lesson).start() 
                    match_subject = one_lesson[match_firsct_close_sc + 1 : match_search_index_start_subject].strip()
                    match_classroom = re.findall(find_classroom, one_lesson)
                    groups_timetable.append({
                        "group": gr_num,
                        "week" : match_week,
                        "day" : timetable.cell(row=j, column=day_column).value,
                        "time": timetable.cell(row=j, column=time_column).value,
                        "subject": match_subject,
                        "teacher": str(match_teacher).replace('[', '').replace(']', '').replace('\'', ''),
                        "classroom": str(match_classroom).replace('[', '').replace(']', '').replace('\'', ''),
                        "address": str(match_address).replace('[', '').replace(']', '').replace('\'', '')
                    })

    return groups_timetable


    # for i in range(time_column + 1, end + 1):
    #     a = timetable.cell(row=groups_row, column=i).value.split(' ')[0]
    #     if(group == a):
    #         group_cell = i; break
        
    # for i in range(groups_row + 1, bottom + 1):
    #     if(timetable.cell(row=i, column=group_cell).value is not None):
    #         if(timetable.cell(row=i, column=day_column).value != timetable.cell(row=i, column=time_column).value != timetable.cell(row=i, column=group_cell).value):
    #             group_timetable.append([timetable.cell(row=i, column=day_column).value, timetable.cell(row=i, column=time_column).value, timetable.cell(row=i, column=group_cell).value])

    # new_subject = []
    # for i in range(len(group_timetable)):
    #     new_subject = group_timetable[i][2]
    #     if(new_subject[0] == "н/н"):

    #         group_timetable_even.append({
    #             "day": group_timetable[i][0],
    #             "time": group_timetable[i][1],
    #             "subject": new_subject[]
    #         })


    # for i in range(len(group_timetable)):
    #     match_week = re.findall(find_week, group_timetable[i][2])
    #     match_teacher = re.findall(find_teacher, group_timetable[i][2])
    #     match_search_teacher = re.search(find_teacher, group_timetable[i][2])
    #     match_address = re.findall(find_street_or_build, group_timetable[i][2])
    #     match_search_address = re.search(find_classroom, group_timetable[i][2])
    #     match_search_index_start_subject = match_search_address.start() if match_search_teacher is None else match_search_teacher.start()
    #     match_firsct_close_sc = re.search(r'\)', group_timetable[i][2]).start() 
    #     match_subject = group_timetable[i][2][match_firsct_close_sc + 1 : match_search_index_start_subject].strip()
    #     match_classroom = re.findall(find_classroom, group_timetable[i][2])
    #     if(match_week == 'н/н'):
    #         group_timetable_even.append({
    #             "day": group_timetable[i][0],
    #             "time": group_timetable[i][1],
    #             "subject": match_subject,
    #             "teacher": match_teacher,
    #             "classroom": match_classroom,
    #             "address": match_address
    #         })
    #     elif(match_week == 'ч/н'):
    #         group_timetable_odd.append({
    #             "day": group_timetable[i][0],
    #             "time": group_timetable[i][1],
    #             "subject": match_subject,
    #             "teacher": match_teacher,
    #             "classroom": match_classroom,
    #             "address": match_address
    #         })
    #     else:
    #         group_timetable_even.append({
    #             "day": group_timetable[i][0],
    #             "time": group_timetable[i][1],
    #             "subject": match_subject,
    #             "teacher": match_teacher,
    #             "classroom": match_classroom,
    #             "address": match_address
    #         })
    #         group_timetable_odd.append({
    #             "day": group_timetable[i][0],
    #             "time": group_timetable[i][1],
    #             "subject": match_subject,
    #             "teacher": match_teacher,
    #             "classroom": match_classroom,
    #             "address": match_address
    #         })

    # return group_timetable_even, group_timetable_odd

# if __name__ == "__main__":
#     print(get_groupp())